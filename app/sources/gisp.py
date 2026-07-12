from __future__ import annotations

import asyncio
import os
import re
import sqlite3
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path
from uuid import uuid4

import httpx
from openpyxl import load_workbook

GISP_ACTIVE_REGISTRY_URL = "https://gisp.gov.ru/pp719v2/mptapp/view/dl/production_res_valid_only/"
GISP_REGISTRY_PAGE_URL = "https://gisp.gov.ru/goods/#/"


class GispRegistryUnavailable(RuntimeError):
    pass


@dataclass(frozen=True)
class GispProduct:
    registry_number: str
    manufacturer: str
    product_name: str | None = None


class GispRegistry:
    def __init__(
        self,
        cache_dir: Path,
        registry_url: str = GISP_ACTIVE_REGISTRY_URL,
        refresh_after: timedelta = timedelta(hours=24),
        timeout_seconds: float = 120.0,
    ) -> None:
        self.cache_dir = cache_dir
        self.registry_url = registry_url
        self.refresh_after = refresh_after
        self.timeout_seconds = timeout_seconds
        self._refresh_lock = asyncio.Lock()

    async def lookup(self, registry_number: str) -> GispProduct | None:
        normalized = normalize_registry_number(registry_number)
        if not normalized:
            return None
        try:
            await self._ensure_index()
            return await asyncio.to_thread(self._lookup_sync, normalized)
        except Exception as exc:
            raise GispRegistryUnavailable(str(exc)) from exc

    async def _ensure_index(self) -> None:
        index_path = self._index_path
        if _is_fresh(index_path, self.refresh_after):
            return
        async with self._refresh_lock:
            if _is_fresh(index_path, self.refresh_after):
                return
            self.cache_dir.mkdir(parents=True, exist_ok=True)
            token = uuid4().hex
            workbook_path = self.cache_dir / f"gisp-registry-{token}.xlsx"
            temporary_index = self.cache_dir / f"gisp-registry-{token}.sqlite3"
            try:
                await self._download(workbook_path)
                await asyncio.to_thread(build_registry_index, workbook_path, temporary_index)
                os.replace(temporary_index, index_path)
            finally:
                workbook_path.unlink(missing_ok=True)
                temporary_index.unlink(missing_ok=True)

    async def _download(self, target: Path) -> None:
        timeout = httpx.Timeout(self.timeout_seconds)
        async with (
            httpx.AsyncClient(timeout=timeout, follow_redirects=True, trust_env=False) as client,
            client.stream("GET", self.registry_url) as response,
        ):
            response.raise_for_status()
            with target.open("wb") as output:
                async for chunk in response.aiter_bytes():
                    output.write(chunk)

    def _lookup_sync(self, normalized_number: str) -> GispProduct | None:
        with sqlite3.connect(self._index_path) as connection:
            row = connection.execute(
                "SELECT registry_number, manufacturer, product_name "
                "FROM products WHERE normalized_number = ? LIMIT 1",
                (normalized_number,),
            ).fetchone()
        if row is None:
            return None
        return GispProduct(registry_number=row[0], manufacturer=row[1], product_name=row[2])

    @property
    def _index_path(self) -> Path:
        return self.cache_dir / "gisp-active-products.sqlite3"


def normalize_registry_number(value: object) -> str:
    return re.sub(r"[^0-9A-ZА-Я]", "", str(value or "").upper().replace("Ё", "Е"))


def build_registry_index(workbook_path: Path, index_path: Path) -> None:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers: dict[str, int] | None = None
    with sqlite3.connect(index_path) as connection:
        connection.execute(
            "CREATE TABLE products ("
            "normalized_number TEXT PRIMARY KEY, registry_number TEXT NOT NULL, "
            "manufacturer TEXT NOT NULL, product_name TEXT)"
        )
        batch: list[tuple[str, str, str, str | None]] = []
        for row in worksheet.iter_rows(values_only=True):
            if headers is None:
                candidate = {str(value).strip(): index for index, value in enumerate(row) if value}
                if "Предприятие" in candidate and "Реестровый номер" in candidate:
                    headers = candidate
                continue
            manufacturer = _cell(row, headers["Предприятие"])
            registry_number = _cell(row, headers["Реестровый номер"])
            if not manufacturer or not registry_number:
                continue
            normalized = normalize_registry_number(registry_number)
            if not normalized:
                continue
            product_name_index = headers.get("Наименование продукции")
            product_name = _cell(row, product_name_index) if product_name_index is not None else None
            batch.append((normalized, registry_number, manufacturer, product_name))
            if len(batch) >= 5000:
                _insert_batch(connection, batch)
                batch.clear()
        if headers is None:
            raise ValueError("GISP registry headers were not found")
        _insert_batch(connection, batch)
        connection.commit()
    workbook.close()


def _insert_batch(connection: sqlite3.Connection, rows: list[tuple[str, str, str, str | None]]) -> None:
    connection.executemany(
        "INSERT OR REPLACE INTO products "
        "(normalized_number, registry_number, manufacturer, product_name) VALUES (?, ?, ?, ?)",
        rows,
    )


def _cell(row: tuple[object, ...], index: int) -> str | None:
    if index >= len(row) or row[index] is None:
        return None
    value = str(row[index]).strip()
    return value or None


def _is_fresh(path: Path, refresh_after: timedelta) -> bool:
    if not path.exists():
        return False
    modified = datetime.fromtimestamp(path.stat().st_mtime, tz=UTC)
    return datetime.now(UTC) - modified < refresh_after
