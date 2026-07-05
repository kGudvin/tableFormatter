from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def inspect_xlsx(path: Path, sheet_name: str = "2026 (44)") -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    header_row = 1
    headers = []
    for cell in sheet[header_row]:
        if cell.value is None:
            continue
        headers.append(
            {
                "index": cell.column,
                "cell": cell.coordinate,
                "title": str(cell.value).strip(),
                "fill": cell.fill.fgColor.rgb or str(cell.fill.fgColor.indexed),
            }
        )
    return {
        "sheets": workbook.sheetnames,
        "active_sheet": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "header_row": header_row,
        "headers": headers,
    }

